import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
import sys
import os
import argparse
import uuid

# Constants for layout snapping
X_TOLERANCE = 3.0
Y_TOLERANCE = 4.0

def hex_color(color_tuple):
    if not color_tuple:
        return "FF000000"
    if len(color_tuple) == 3:
        try:
            return f"FF{int(color_tuple[0]*255):02x}{int(color_tuple[1]*255):02x}{int(color_tuple[2]*255):02x}"
        except:
            return "FF000000"
    elif len(color_tuple) == 1:
        try:
            val = int(color_tuple[0]*255)
            return f"FF{val:02x}{val:02x}{val:02x}"
        except:
            return "FF000000"
    return "FF000000"

def get_grid_boundaries(elements, tolerance):
    bounds = set()
    for el in elements:
        bounds.add(el[0]) # start
        bounds.add(el[1]) # end
    
    sorted_bounds = sorted(list(bounds))
    if not sorted_bounds:
        return []
    
    merged = [sorted_bounds[0]]
    for b in sorted_bounds[1:]:
        if b - merged[-1] > tolerance:
            merged.append(b)
    return merged

def get_start_index(value, boundaries, tol):
    n = len(boundaries) - 1
    for i in range(len(boundaries)):
        if abs(value - boundaries[i]) <= tol:
            return i + 1 if i < n else n
    for i in range(1, n + 1):
        if value <= boundaries[i]:
            return i
    return n

def get_end_index(value, boundaries, tol):
    n = len(boundaries) - 1
    for i in range(len(boundaries)):
        if abs(value - boundaries[i]) <= tol:
            return max(1, i)
    for i in range(1, n + 1):
        if value <= boundaries[i]:
            return i
    return n

def points_to_excel_width(points):
    width = points / 5.25
    return max(0.1, width)

def get_text_blocks(words):
    if not words: return []
    
    # Deduplicate exact overlapping words (common in OCR PDFs with invisible text layers)
    unique_words = []
    seen = set()
    for w in words:
        # Round to 1 decimal place to catch near-exact overlaps
        coord = (w['text'], round(w['x0'], 1), round(w['top'], 1))
        if coord not in seen:
            seen.add(coord)
            unique_words.append(w)
            
    words = sorted(unique_words, key=lambda w: (w['top'], w['x0']))
    
    lines = []
    current_line = []
    current_top = words[0]['top']
    current_bottom = words[0]['bottom']
    
    for w in words:
        if w['top'] <= current_bottom + 1 and w['bottom'] >= current_top - 1:
            current_line.append(w)
            current_top = min(current_top, w['top'])
            current_bottom = max(current_bottom, w['bottom'])
        else:
            lines.append(current_line)
            current_line = [w]
            current_top = w['top']
            current_bottom = w['bottom']
    if current_line:
        lines.append(current_line)
        
    blocks = []
    for line in lines:
        line = sorted(line, key=lambda w: w['x0'])
        current_block = [line[0]]
        for w in line[1:]:
            prev = current_block[-1]
            if w['x0'] - prev['x1'] < 10:
                current_block.append(w)
            else:
                blocks.append(current_block)
                current_block = [w]
        if current_block:
            blocks.append(current_block)
            
    final_blocks = []
    for b in blocks:
        text = " ".join([w['text'] for w in b])
        x0 = min(w['x0'] for w in b)
        top = min(w['top'] for w in b)
        x1 = max(w['x1'] for w in b)
        bottom = max(w['bottom'] for w in b)
        w0 = b[0]
        final_blocks.append({
            'text': text,
            'x0': x0, 'top': top, 'x1': x1, 'bottom': bottom,
            'fontname': w0.get('fontname', 'Calibri'),
            'size': w0.get('size', 11),
            'color': w0.get('non_stroking_color', (0,0,0))
        })
    return final_blocks

def extract_graphics(page):
    graphics = []
    for rect in page.rects:
        graphics.append({
            'type': 'rect',
            'x0': rect['x0'], 'top': rect['top'], 'x1': rect['x1'], 'bottom': rect['bottom'],
            'fill': rect.get('non_stroking_color'),
            'stroke': rect.get('stroking_color'),
            'linewidth': rect.get('linewidth', 1)
        })
    for line in page.lines:
        graphics.append({
            'type': 'line',
            'x0': min(line['x0'], line['x1']),
            'top': min(line['top'], line['bottom']),
            'x1': max(line['x0'], line['x1']),
            'bottom': max(line['top'], line['bottom']),
            'stroke': line.get('stroking_color'),
            'linewidth': line.get('linewidth', 1)
        })
    for curve in page.curves:
        graphics.append({
            'type': 'rect',
            'x0': curve['x0'], 'top': curve['top'], 'x1': curve['x1'], 'bottom': curve['bottom'],
            'fill': curve.get('non_stroking_color'),
            'stroke': curve.get('stroking_color'),
            'linewidth': curve.get('linewidth', 1)
        })
    return graphics

def render_grid_to_sheet(ws, blocks, graphics, images_data):
    x_elements = []
    y_elements = []
    
    for b in blocks:
        x_elements.append((b['x0'], b['x1']))
        y_elements.append((b['top'], b['bottom']))
        
    for g in graphics:
        if g['x1'] - g['x0'] > 0.5:
            x_elements.append((g['x0'], g['x1']))
        if g['bottom'] - g['top'] > 0.5:
            y_elements.append((g['top'], g['bottom']))
            
    for img_d in images_data:
        b = img_d['bbox']
        x_elements.append((b[0], b[2]))
        y_elements.append((b[1], b[3]))
    
    col_bounds = get_grid_boundaries(x_elements, X_TOLERANCE)
    row_bounds = get_grid_boundaries(y_elements, Y_TOLERANCE)
    
    if not col_bounds or not row_bounds:
        ws["A1"] = "Blank Page"
        return
        
    if col_bounds[0] > 0: col_bounds.insert(0, 0)
    if row_bounds[0] > 0: row_bounds.insert(0, 0)
    
    for i in range(len(col_bounds) - 1):
        width_pts = col_bounds[i+1] - col_bounds[i]
        ws.column_dimensions[get_column_letter(i+1)].width = points_to_excel_width(width_pts)
        
    for i in range(len(row_bounds) - 1):
        height_pts = row_bounds[i+1] - row_bounds[i]
        ws.row_dimensions[i+1].height = max(1, height_pts)
        
    def map_to_grid(x0, top, x1, bottom):
        sc = get_start_index(x0, col_bounds, X_TOLERANCE)
        ec = max(sc, get_end_index(x1, col_bounds, X_TOLERANCE))
        sr = get_start_index(top, row_bounds, Y_TOLERANCE)
        er = max(sr, get_end_index(bottom, row_bounds, Y_TOLERANCE))
        return sc, sr, ec, er
    
    for g in graphics:
        sc, sr, ec, er = map_to_grid(g['x0'], g['top'], g['x1'], g['bottom'])
        is_line = (ec - sc == 0) or (er - sr == 0) or g['type'] == 'line'
        
        if not is_line:
            if g.get('fill'):
                fill_color = hex_color(g['fill'])
                if fill_color not in ("ffffff", "FFFFFFFF"):
                    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    for row in range(sr, er + 1):
                        for col in range(sc, ec + 1):
                            ws.cell(row=row, column=col).fill = fill
        
        if g.get('stroke'):
            border_color = hex_color(g['stroke'])
            side = Side(style='thin', color=border_color)
            
            if is_line:
                if er == sr:
                    cell_top = row_bounds[sr-1]
                    cell_bottom = row_bounds[sr]
                    line_y = (g['top'] + g['bottom']) / 2
                    if abs(line_y - cell_top) < abs(line_y - cell_bottom):
                        for col in range(sc, ec + 1):
                            c = ws.cell(row=sr, column=col)
                            c.border = Border(top=side, bottom=c.border.bottom, left=c.border.left, right=c.border.right)
                    else:
                        for col in range(sc, ec + 1):
                            c = ws.cell(row=sr, column=col)
                            c.border = Border(bottom=side, top=c.border.top, left=c.border.left, right=c.border.right)
                elif ec == sc:
                    cell_left = col_bounds[sc-1]
                    cell_right = col_bounds[sc]
                    line_x = (g['x0'] + g['x1']) / 2
                    if abs(line_x - cell_left) < abs(line_x - cell_right):
                        for row in range(sr, er + 1):
                            c = ws.cell(row=row, column=sc)
                            c.border = Border(left=side, right=c.border.right, top=c.border.top, bottom=c.border.bottom)
                    else:
                        for row in range(sr, er + 1):
                            c = ws.cell(row=row, column=sc)
                            c.border = Border(right=side, left=c.border.left, top=c.border.top, bottom=c.border.bottom)
            else:
                for col in range(sc, ec + 1):
                    c = ws.cell(row=sr, column=col)
                    c.border = Border(top=side, bottom=c.border.bottom, left=c.border.left, right=c.border.right)
                    c = ws.cell(row=er, column=col)
                    c.border = Border(bottom=side, top=c.border.top, left=c.border.left, right=c.border.right)
                for row in range(sr, er + 1):
                    c = ws.cell(row=row, column=sc)
                    c.border = Border(left=side, right=c.border.right, top=c.border.top, bottom=c.border.bottom)
                    c = ws.cell(row=row, column=ec)
                    c.border = Border(right=side, left=c.border.left, top=c.border.top, bottom=c.border.bottom)

    for b in blocks:
        sc, sr, ec, er = map_to_grid(b['x0'], b['top'], b['x1'], b['bottom'])
        
        if sc != ec or sr != er:
            try:
                ws.merge_cells(start_row=sr, start_column=sc, end_row=er, end_column=ec)
            except Exception:
                pass
                
        cell = ws.cell(row=sr, column=sc)
        
        if cell.value:
            cell.value = str(cell.value) + " " + b['text']
        else:
            cell.value = b['text']
            
        font_name_str = str(b['fontname'])
        font_size = b['size']
        is_bold = 'Bold' in font_name_str or 'Black' in font_name_str
        is_italic = 'Italic' in font_name_str
        font_color = hex_color(b['color'])
        
        cell.font = Font(
            name='Calibri',
            size=font_size if font_size < 40 else 14,
            bold=is_bold,
            italic=is_italic,
            color=font_color
        )
        
        cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')

    for img_d in images_data:
        sc, sr, ec, er = map_to_grid(*img_d['bbox'])
        try:
            img_excel = OpenpyxlImage(img_d['path'])
            cell_anchor = f"{get_column_letter(sc)}{sr}"
            width_pts = img_d['bbox'][2] - img_d['bbox'][0]
            height_pts = img_d['bbox'][3] - img_d['bbox'][1]
            img_excel.width = int(width_pts * 1.33)
            img_excel.height = int(height_pts * 1.33)
            ws.add_image(img_excel, cell_anchor)
        except Exception as e:
            print(f"Error placing image: {e}")


def extract_words_with_ocr(page, ocr=False):
    words = page.extract_words(keep_blank_chars=True, extra_attrs=["fontname", "size", "non_stroking_color"])
    if (len(words) < 20 or ocr):
        try:
            import pytesseract
            from pytesseract import Output
            
            # Point to default Tesseract location on Windows if not in PATH
            tess_path = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
            if os.path.exists(tess_path):
                pytesseract.pytesseract.tesseract_cmd = tess_path
                
            img_obj = page.to_image(resolution=150)
            data = pytesseract.image_to_data(img_obj.original, output_type=Output.DICT)
            for i in range(len(data['text'])):
                if int(data['conf'][i]) > 30 and data['text'][i].strip():
                    x0 = data['left'][i] / 150 * 72
                    top = data['top'][i] / 150 * 72
                    x1 = x0 + data['width'][i] / 150 * 72
                    bottom = top + data['height'][i] / 150 * 72
                    words.append({
                        'text': data['text'][i],
                        'x0': x0, 'top': top, 'x1': x1, 'bottom': bottom,
                        'fontname': 'Calibri',
                        'size': 11,
                        'non_stroking_color': (0,0,0)
                    })
        except Exception as e:
            print(f"OCR warning: {e}")
    return words

def process_pdf_to_excel(input_path, output_path, layout="multiple_sheets", ocr=False):
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
        
    all_temp_images = []
    try:
        with pdfplumber.open(input_path) as pdf:
            if not pdf.pages:
                raise ValueError("PDF has no pages")
            
            temp_dir = os.path.dirname(output_path)
            
            if layout == "one_sheet":
                ws = wb.create_sheet(title="Converted Document")
                ws.sheet_view.showGridLines = False
                ws.page_setup.paperSize = ws.PAPERSIZE_A4
                
                global_blocks = []
                global_graphics = []
                global_images = []
                y_offset = 0
                
                for page_idx, page in enumerate(pdf.pages):
                    words = extract_words_with_ocr(page, ocr)
                    blocks = get_text_blocks(words)
                    graphics = extract_graphics(page)
                    images_data = []
                    
                    if len(words) < 20 and not ocr:
                        try:
                            temp_page_img = os.path.join(temp_dir, f"temp_page_full_{uuid.uuid4().hex}.png")
                            img_obj = page.to_image(resolution=150)
                            img_obj.save(temp_page_img, format="PNG")
                            images_data.append({'bbox': (0, 0, page.width, page.height), 'path': temp_page_img})
                            all_temp_images.append(temp_page_img)
                        except Exception as e:
                            print(f"Warning: Could not render page {page_idx} as image: {e}")
                            
                    for idx, img in enumerate(page.images):
                        try:
                            bbox = (img['x0'], img['top'], img['x1'], img['bottom'])
                            stream = img.get('stream')
                            if stream:
                                temp_img_path = os.path.join(temp_dir, f"temp_img_{uuid.uuid4().hex}.png")
                                cropped = page.within_bbox(bbox)
                                img_obj = cropped.to_image(resolution=150)
                                img_obj.save(temp_img_path, format="PNG")
                                images_data.append({'bbox': bbox, 'path': temp_img_path})
                                all_temp_images.append(temp_img_path)
                        except Exception as e:
                            print(f"Warning: Could not extract image {idx} on page {page_idx}: {e}")
                    
                    for b in blocks:
                        b['top'] += y_offset
                        b['bottom'] += y_offset
                        global_blocks.append(b)
                        
                    for g in graphics:
                        g['top'] += y_offset
                        g['bottom'] += y_offset
                        global_graphics.append(g)
                        
                    for img_d in images_data:
                        b = img_d['bbox']
                        img_d['bbox'] = (b[0], b[1] + y_offset, b[2], b[3] + y_offset)
                        global_images.append(img_d)
                        
                    y_offset += page.height + 20 # 20 points gap between pages
                
                render_grid_to_sheet(ws, global_blocks, global_graphics, global_images)
                
            else:
                for page_idx, page in enumerate(pdf.pages):
                    ws = wb.create_sheet(title=f"Page_{page_idx + 1}")
                    ws.sheet_view.showGridLines = False
                    ws.page_setup.paperSize = ws.PAPERSIZE_A4
                    
                    words = extract_words_with_ocr(page, ocr)
                    blocks = get_text_blocks(words)
                    graphics = extract_graphics(page)
                    images_data = []
                    
                    if len(words) < 20 and not ocr:
                        try:
                            temp_page_img = os.path.join(temp_dir, f"temp_page_full_{uuid.uuid4().hex}.png")
                            img_obj = page.to_image(resolution=150)
                            img_obj.save(temp_page_img, format="PNG")
                            images_data.append({'bbox': (0, 0, page.width, page.height), 'path': temp_page_img})
                            all_temp_images.append(temp_page_img)
                        except Exception as e:
                            print(f"Warning: Could not render page {page_idx} as image: {e}")
                            
                    for idx, img in enumerate(page.images):
                        try:
                            bbox = (img['x0'], img['top'], img['x1'], img['bottom'])
                            stream = img.get('stream')
                            if stream:
                                temp_img_path = os.path.join(temp_dir, f"temp_img_{uuid.uuid4().hex}.png")
                                cropped = page.within_bbox(bbox)
                                img_obj = cropped.to_image(resolution=150)
                                img_obj.save(temp_img_path, format="PNG")
                                images_data.append({'bbox': bbox, 'path': temp_img_path})
                                all_temp_images.append(temp_img_path)
                        except Exception as e:
                            print(f"Warning: Could not extract image {idx} on page {page_idx}: {e}")
                            
                    render_grid_to_sheet(ws, blocks, graphics, images_data)

    except Exception as e:
        print(f"Error processing PDF: {str(e)}")
        if not wb.sheetnames:
            ws = wb.create_sheet("Error")
            ws["A1"] = f"Conversion failed: {str(e)}"
            
    wb.save(output_path)
    
    # Cleanup all images after saving the workbook
    for img_path in all_temp_images:
        if os.path.exists(img_path):
            os.remove(img_path)
    print("SUCCESS")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Convert PDF to Excel.")
    parser.add_argument("input_pdf", help="Input PDF file")
    parser.add_argument("output_xlsx", help="Output XLSX file")
    parser.add_argument("--layout", default="multiple_sheets", choices=["one_sheet", "multiple_sheets"], help="Layout format")
    parser.add_argument("--ocr", default="false", choices=["true", "false"], help="Use OCR")
    args = parser.parse_args()
    
    if not os.path.exists(args.input_pdf):
        print(f"ERROR: File not found {args.input_pdf}")
        sys.exit(1)
        
    process_pdf_to_excel(args.input_pdf, args.output_xlsx, layout=args.layout, ocr=(args.ocr == "true"))
