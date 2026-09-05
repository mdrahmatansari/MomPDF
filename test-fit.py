import openpyxl
import sys

def auto_fit(input_path, output_path):
    wb = openpyxl.load_workbook(input_path)
    for sheet in wb.worksheets:
        # Enable fit to page property
        if sheet.sheet_properties.pageSetUpPr is None:
            sheet.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
        else:
            sheet.sheet_properties.pageSetUpPr.fitToPage = True
            
        # Fit all columns to 1 page wide, but let rows span multiple pages
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        
    wb.save(output_path)

if __name__ == "__main__":
    if len(sys.argv) > 2:
        auto_fit(sys.argv[1], sys.argv[2])
    else:
        # create a dummy file to test
        wb = openpyxl.Workbook()
        ws = wb.active
        for i in range(1, 100):
            for j in range(1, 30):
                ws.cell(row=i, column=j, value=f"Cell {i},{j}")
        wb.save("test_input.xlsx")
        auto_fit("test_input.xlsx", "test_fitted.xlsx")
        print("Done")
